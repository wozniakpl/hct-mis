import openpyxl
import copy
from django.core.files import File
from openpyxl.utils import get_column_letter
from django.db.models import Min, Max, Sum, Q, Count
from django.contrib.postgres.aggregates.general import ArrayAgg
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from tempfile import NamedTemporaryFile

from hct_mis_api.apps.core.models import AdminArea
from hct_mis_api.apps.core.utils import encode_id_base64
from hct_mis_api.apps.reporting.models import DashboardReport
from hct_mis_api.apps.household.models import Individual, Household, ACTIVE
from hct_mis_api.apps.program.models import CashPlanPaymentVerification, CashPlan, Program
from hct_mis_api.apps.payment.models import PaymentRecord, PaymentVerification
from hct_mis_api.apps.core.utils import decode_id_string
from hct_mis_api.apps.account.models import User


class GenerateDashboardReportContentHelpers:
    @staticmethod
    def _to_values_list(instances, field_name: str) -> str:
        values_list = list(instances.values_list(field_name, flat=True))
        return ", ".join([str(value) for value in values_list])

    @staticmethod
    def _format_date(date) -> str:
        if not date:
            return ""
        return date.strftime("%Y-%m-%d")

    @staticmethod
    def _is_report_global(report):
        return report.business_area.slug == "global"

    @staticmethod
    def get_beneficiaries(report: DashboardReport):
        # TODO: implement
        return []

    @classmethod
    def format_beneficiaries(self, instance, is_hq: bool) -> tuple:
        # TODO: implement
        return ()


class GenerateDashboardReportService:
    HQ = 1
    COUNTRY = 2
    SHARED = 3
    HEADERS = {
        DashboardReport.BENEFICIARIES_REACHED: {
            HQ: ("business area", "country"),
            COUNTRY: ("business area", "programme"),
            SHARED: ("households reached", "individuals reached", "children reached"),
        },
        DashboardReport.TOTAL_TRANSFERRED_BY_ADMIN_AREA: {
            HQ: (),
            COUNTRY: (
                "Admin Level 2",
                "Admin Code",
                "Total tranferred (USD)",
                "Households reached",
                "Female 0-5 Reached",
                "Female 6-11 Reached",
                "Female 12-17 Reached",
                "Female 18-59 Reached",
                "Female 60+ Reached",
                "Male 0-5 Reached",
                "Male 6-11 Reached",
                "Male 12-17 Reached",
                "Male 18-59 Reached",
                "Male 60+ Reached",
            ),
            SHARED: (),
        },
        DashboardReport.PAYMENT_VERIFICATION: {
            HQ: (),
            COUNTRY: (),
            SHARED: (
                "business area",
                "country",
                "programme",
                "cash plan verifications",
                "Households Contacted",
                "average sampling",
                "Received",
                "Not Received",
                "Received with issues",
                "Not Responded",
            ),
        },
        DashboardReport.GRIEVANCES_AND_FEEDBACK: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: (
                "grievance tickets",
                "feedback tickets",
                "resolved tickets",
                "Unresolved >30 days",
                "Unresolved >60 days",
                "open sensitive grievances",
            ),
        },
        DashboardReport.TOTAL_TRANSFERRED_BY_COUNTRY: {
            HQ: ("business area", "country", "actual cash transferred", "actual voucher transferred"),
            COUNTRY: (),
            SHARED: (),
        },
        DashboardReport.PROGRAMS: {
            HQ: (),
            COUNTRY: (),
            SHARED: (
                "business area",
                "country",
                "programme",
                "sector",
                "cash+",
                "frequency",
                "unsuccessful payment",
                "January cash",
                "January voucher",
                "February cash",
                "February voucher",
                "March cash",
                "March voucher",
                "April cash",
                "April voucher",
                "May cash",
                "May voucher",
                "June cash",
                "June voucher",
                "July cash",
                "July voucher",
                "August cash",
                "August voucher",
                "September cash",
                "September voucher",
                "October cash",
                "October voucher",
                "November cash",
                "November voucher",
                "December cash",
                "December voucher",
            ),
        },
        DashboardReport.VOLUME_BY_DELIVERY_MECHANISM: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: (
                "Cash in envelope",
                "cash by FSP",
                "deposit to card",
                "mobile money",
                "voucher",
                "e-voucher",
            ),
        },
        DashboardReport.INDIVIDUALS_REACHED: {
            HQ: (
                "business area",
                "country",
            ),
            COUNTRY: (
                "business area",
                "programme",
            ),
            SHARED: (
                "females 0-5 reached",
                "females 0-5 w/ disability reached",
                "females 6-11 reached",
                "females 6-11 w/ disability reached",
                "females 12-17 reached",
                "females 12-17 w/ disability reached",
                "females 18-59 reached",
                "females 18-59 w/ disability reached",
                "females 60+ reached",
                "females 60+ w/ disability reached",
                "males 0-5 reached",
                "males 0-5 w/ disability reached",
                "males 6-11 reached",
                "males 6-11 w/ disability reached",
                "males 12-17 reached",
                "males 12-17 w/ disability reached",
                "males 18-59 reached",
                "males 18-59 w/ disability reached",
                "males 60+ reached",
                "males 60+ w/ disability reached",
            ),
        },
    }
    ROW_CONTENT_METHODS = {
        DashboardReport.BENEFICIARIES_REACHED: (
            GenerateDashboardReportContentHelpers.get_beneficiaries,
            GenerateDashboardReportContentHelpers.format_beneficiaries,
        ),
        # TODO: add the rest of the methods
    }
    META_HEADERS = ("report type", "creation date", "created by", "business area", "report year")
    META_SHEET = "Meta data"
    MAX_COL_WIDTH = 75

    def __init__(self, report: DashboardReport):
        self.report = report
        self.report_types = report.report_type
        self.business_area = report.business_area
        # TODO check if this is best way to determin if global
        self.hq_or_country = self.HQ if report.business_area.slug == "global" else self.COUNTRY

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_meta = wb.active
        ws_meta.title = self.META_SHEET
        self.wb = wb
        self.ws_meta = ws_meta
        return wb

    def _format_meta_tab(self):
        self.ws_meta.append(self.META_HEADERS)
        info_row = (
            self._report_types_to_joined_str(),
            self._format_date(self.report.created_at),
            self._format_user_name(self.report.created_by),
            self.business_area.name,
            str(self.report.year),
        )
        self.ws_meta.append(info_row)

    def _add_headers(self, active_sheet, report_type) -> int:
        headers_row = self.HEADERS[report_type][self.hq_or_country] + self.HEADERS[report_type][self.SHARED]
        active_sheet.append(headers_row)
        return len(headers_row)

    def _add_rows(self, active_sheet, report_type):
        get_row_methods = self.ROW_CONTENT_METHODS[report_type]
        all_instances = get_row_methods[0](self.report)
        for instance in all_instances:
            row = get_row_methods[1](instance, self.hq_or_country == self.HQ)
            str_row = self._stringify_all_values(row)
            active_sheet.append(str_row)

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._format_meta_tab()
        self._adjust_column_width_from_col(self.ws_meta, 1, 5, 1)

        # loop through all selected report types and add sheet for each
        for report_type in self.report_types:
            print("IN FOR LOOP", report_type)
            sheet_title = self._report_type_to_str(report_type)
            print("SHEET TITLE", sheet_title)
            active_sheet = self.wb.create_sheet(sheet_title)
            print("CREATED ACTIVE SHEET")
            number_of_columns = self._add_headers(active_sheet, report_type)
            print("ADDED HEADERS")
            self._add_rows()
            print("ADDED ROWS")
            self._adjust_column_width_from_col(active_sheet, 1, number_of_columns, 1)
            print("ADJUTED WIDTH")
        return self.wb

    def generate_report(self):
        try:
            self.generate_workbook()
            print(self.wb)
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                self.report.file.save(
                    f"{self._report_types_to_joined_str()}-{GenerateDashboardReportContentHelpers._format_date(self.report.created_at)}.xlsx",
                    File(tmp),
                    save=False,
                )
                self.report.status = DashboardReport.COMPLETED
        except Exception as e:
            print("ERROR", e)
            self.report.status = DashboardReport.FAILED
        self.report.save()

        if self.report.file:
            self._send_email()

    def _send_email(self):
        pass
        # context = {
        #     "report_type": self._report_type_to_str(),
        #     "created_at": GenerateReportContentHelpers._format_date(self.report.created_at),
        #     "report_url": f'https://{settings.FRONTEND_HOST}/{self.business_area.slug}/reporting/{encode_id_base64(self.report.id, "Report")}',
        # }
        # text_body = render_to_string("report.txt", context=context)
        # html_body = render_to_string("report.html", context=context)
        # msg = EmailMultiAlternatives(
        #     subject="HOPE report generated",
        #     from_email=settings.EMAIL_HOST_USER,
        #     to=[self.report.created_by.email],
        #     body=text_body,
        # )
        # msg.attach_alternative(html_body, "text/html")
        # msg.send()

    def _adjust_column_width_from_col(self, ws, min_col, max_col, min_row):

        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):

            for cell in col:
                value = cell.value

                if value is not None:

                    if isinstance(value, str) is False:
                        value = str(value)

                    if len(value) > GenerateDashboardReportService.MAX_COL_WIDTH:
                        alignment = copy.copy(cell.alignment)
                        alignment.wrapText = True
                        cell.alignment = alignment

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            value = (
                GenerateDashboardReportService.MAX_COL_WIDTH
                if value > GenerateDashboardReportService.MAX_COL_WIDTH
                else value
            )
            ws.column_dimensions[col_name].width = value

    def _report_type_to_str(self, report_type) -> str:
        return str([name for value, name in DashboardReport.REPORT_TYPES if value == report_type][0])

    def _report_types_to_joined_str(self) -> str:
        return ", ".join([self._report_type_to_str(report_type) for report_type in self.report_types])

    def _stringify_all_values(self, row: tuple) -> tuple:
        str_row = []
        for value in row:
            str_row.append(str(value if value is not None else ""))
        return tuple(str_row)

    def _format_date(self, date) -> str:
        if not date:
            return ""
        return date.strftime("%Y-%m-%d")

    def _format_user_name(self, user: User) -> str:
        return (
            f"{user.first_name} {user.last_name}" if user.first_name or user.last_name else user.email or user.username
        )
