import openpyxl
from django.core.files import File
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from tempfile import NamedTemporaryFile

from hct_mis_api.apps.core.models import AdminArea
from hct_mis_api.apps.reporting.models import Report
from hct_mis_api.apps.household.models import Individual, Household, ACTIVE
from hct_mis_api.apps.program.models import CashPlanPaymentVerification, CashPlan, Program
from hct_mis_api.apps.payment.models import PaymentRecord, PaymentVerification
from hct_mis_api.apps.core.utils import decode_id_string


class GenerateReportContentHelpers:
    @staticmethod
    def _get_individuals(report: Report):
        filter_vars = {
            "household__business_area": report.business_area,
            "status": ACTIVE,
            "last_registration_date__gte": report.date_from,
            "last_registration_date__lte": report.date_to,
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        return Individual.objects.filter(**filter_vars)

    @classmethod
    def _format_individual_row(self, individual: Individual) -> tuple:

        return (
            individual.household.id,
            individual.household.country_origin.name if individual.household.country_origin else "",
            individual.household.admin_area.title if individual.household.admin_area else "",
            individual.birth_date,
            individual.estimated_birth_date,
            individual.sex,
            individual.marital_status,
            individual.disability,
            individual.observed_disability,
            individual.comms_disability,
            individual.hearing_disability,
            individual.memory_disability,
            individual.physical_disability,
            individual.seeing_disability,
            individual.selfcare_disability,
            individual.pregnant,
            individual.relationship,
            self._to_values_list(individual.households_and_roles.all(), "role"),
            individual.work_status,
            individual.sanction_list_possible_match,
            individual.deduplication_batch_status,
            individual.deduplication_golden_record_status,
            individual.deduplication_golden_record_results.get("duplicates", "")
            if individual.deduplication_golden_record_results
            else "",
            individual.deduplication_golden_record_results.get("possible_duplicates", "")
            if individual.deduplication_golden_record_results
            else "",
        )

    @staticmethod
    def _get_households(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "status": ACTIVE,
            "last_registration_date__gte": report.date_from,
            "last_registration_date__lte": report.date_to,
        }
        if report.admin_area.all().exists():
            filter_vars["admin_area__in"] = report.admin_area.all()
        return Household.objects.filter(**filter_vars)

    @classmethod
    def _format_household_row(self, household: Household) -> tuple:
        row = [
            household.id,
            household.country_origin.name if household.country_origin else "",
            household.admin_area.title if household.admin_area else "",
            household.size,
            household.geopoint[0] if household.geopoint else "",
            household.geopoint[1] if household.geopoint else "",
            household.residence_status,
            household.returnee,
            household.status,
            household.village,
            household.female_age_group_0_5_count,
            household.female_age_group_0_5_disabled_count,
            household.female_age_group_6_11_count,
            household.female_age_group_6_11_disabled_count,
            household.female_age_group_12_17_count,
            household.female_age_group_12_17_disabled_count,
            household.female_age_group_18_59_count,
            household.female_age_group_18_59_disabled_count,
            household.female_age_group_60_count,
            household.female_age_group_60_disabled_count,
            household.pregnant_count,
            household.male_age_group_0_5_count,
            household.male_age_group_0_5_disabled_count,
            household.male_age_group_6_11_count,
            household.male_age_group_6_11_disabled_count,
            household.male_age_group_12_17_count,
            household.male_age_group_12_17_disabled_count,
            household.male_age_group_18_59_count,
            household.male_age_group_18_59_disabled_count,
            household.male_age_group_60_count,
            household.male_age_group_60_disabled_count,
            household.first_registration_date,
            household.last_registration_date,
            household.org_name_enumerator,
        ]
        for program in household.programs.all():
            row.append(program.name)
        return tuple(row)

    @staticmethod
    def _get_cash_plan_verifications(report: Report):
        filter_vars = {
            "cash_plan__business_area": report.business_area,
            "completion_date__isnull": False,
            "completion_date__gte": report.date_from,
            "completion_date__lte": report.date_to,
        }
        if report.program:
            filter_vars["cash_plan__program"] = report.program
        return CashPlanPaymentVerification.objects.filter(**filter_vars)

    @staticmethod
    def _map_admin_area_names_from_ids(admin_areas_ids: list) -> str:
        if not admin_areas_ids:
            return ""
        result = []
        for admin_area_id in admin_areas_ids:
            admin_area_id = decode_id_string(admin_area_id)
            admin_area = AdminArea.objects.filter(id=admin_area_id).first()
            if admin_area:
                result.append(admin_area.title)
        return ", ".join(result)

    @classmethod
    def _format_cash_plan_verification_row(self, verification: CashPlanPaymentVerification) -> tuple:
        return (
            verification.cash_plan.ca_id,
            verification.id,
            verification.cash_plan.program.name,
            verification.activation_date,
            verification.status,
            verification.verification_method,
            verification.completion_date,
            verification.sample_size,
            verification.responded_count,
            verification.received_count,
            verification.received_with_problems_count,
            verification.not_received_count,
            verification.sampling,
            verification.sex_filter,
            self._map_admin_area_names_from_ids(verification.excluded_admin_areas_filter),
            verification.age_filter,
        )

    @staticmethod
    def _get_payments(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "delivery_date__gte": report.date_from,
            "delivery_date__lte": report.date_to,
        }
        if report.admin_area.all().exists():
            filter_vars["household__admin_area__in"] = report.admin_area.all()
        return PaymentRecord.objects.filter(**filter_vars)

    @classmethod
    def _format_payment_row(self, payment: PaymentRecord) -> tuple:
        cash_or_voucher = ""
        if payment.delivery_type:
            if payment.delivery_type in [
                PaymentRecord.DELIVERY_TYPE_CASH,
                PaymentRecord.DELIVERY_TYPE_DEPOSIT_TO_CARD,
                PaymentRecord.DELIVERY_TYPE_TRANSFER,
            ]:
                cash_or_voucher = "cash"
            else:
                # TODO: check if this is even an option, I don't see the delivery_type being anything else but above three options
                # but following the spreadsheet here
                cash_or_voucher = "voucher"

        return (
            payment.cash_plan.ca_id if payment.cash_plan else "",
            payment.ca_id,
            payment.status,
            payment.currency,
            payment.delivered_quantity,
            payment.delivery_date,
            payment.delivery_type,
            payment.distribution_modality,
            payment.entitlement_quantity,
            payment.target_population.id,
            payment.target_population.name,
            cash_or_voucher,
        )

    @staticmethod
    def _get_payment_verifications(report: Report):
        filter_vars = {
            "cash_plan_payment_verification__cash_plan__business_area": report.business_area,
            "cash_plan_payment_verification__completion_date__isnull": False,
            "cash_plan_payment_verification__completion_date__gte": report.date_from,
            "cash_plan_payment_verification__completion_date__lte": report.date_to,
        }
        if report.program:
            filter_vars["cash_plan_payment_verification__cash_plan__program"] = report.program
        return PaymentVerification.objects.filter(**filter_vars)

    @classmethod
    def _format_payment_verification_row(self, payment_verification: PaymentVerification) -> tuple:
        return (
            payment_verification.payment_record.ca_id,
            payment_verification.cash_plan_payment_verification.cash_plan.ca_id,
            payment_verification.cash_plan_payment_verification.id,
            payment_verification.cash_plan_payment_verification.completion_date,
            payment_verification.received_amount,
            payment_verification.status,
            payment_verification.status_date,
        )

    @staticmethod
    def _get_cash_plans(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "end_date__gte": report.date_from,
            "end_date__lte": report.date_to,
        }
        if report.program:
            filter_vars["cash_plan__program"] = report.program
        return CashPlan.objects.filter(**filter_vars)

    @classmethod
    def _format_cash_plan_row(self, cash_plan: CashPlan) -> tuple:
        return (
            cash_plan.program.name,
            cash_plan.assistance_measurement,
            cash_plan.assistance_through,
            cash_plan.business_area.id,
            cash_plan.ca_hash_id,
            cash_plan.delivery_type,
            cash_plan.dispersion_date,
            cash_plan.down_payment,
            cash_plan.end_date,
            cash_plan.funds_commitment,
            cash_plan.name,
            cash_plan.program.id,
            cash_plan.start_date,
            cash_plan.status,
            cash_plan.status_date,
            cash_plan.total_delivered_quantity,
            cash_plan.total_entitled_quantity,
            cash_plan.total_entitled_quantity_revised,
            cash_plan.total_persons_covered,
            cash_plan.total_persons_covered_revised,
            cash_plan.total_undelivered_quantity,
            cash_plan.validation_alerts_count,
            cash_plan.verification_status,
            cash_plan.vision_id,
        )

    @staticmethod
    def _get_programs(report: Report):
        filter_vars = {
            "business_area": report.business_area,
            "end_date__gte": report.date_from,
            "end_date__lte": report.date_to,
        }
        return Program.objects.filter(**filter_vars)

    @classmethod
    def _format_program_row(self, program: Program) -> tuple:
        return (
            program.business_area.id,
            program.administrative_areas_of_implementation,
            program.budget,
            program.cash_plus,
            program.description,
            program.end_date,
            program.frequency_of_payments,
            program.id,
            program.name,
            program.population_goal,
            program.scope,
            program.sector,
            program.start_date,
            program.status,
            program.total_number_of_households,
        )

    @staticmethod
    def _get_payments_for_individuals(report: Report):
        # TODO fix this
        # delivery date for timeframe
        return PaymentRecord.objects.none()

    @staticmethod
    def _format_payments_for_individuals_row(self, payment_record: PaymentRecord) -> tuple:
        # TODO: fix this
        return ()

    @staticmethod
    def _to_values_list(instances, field_name: str) -> str:
        values_list = list(instances.values_list(field_name, flat=True))
        return ", ".join([str(value) for value in values_list])

    @staticmethod
    def _sum_values(*values):
        total = 0
        for value in values:
            total = total + value if value else total
        return total


class GenerateReportService:
    HEADERS = {
        Report.INDIVIDUALS: (
            "household id",  # 8e8ea94a-2ca5-4b76-b055-e098bc24eee8
            "country of origin",  # South Sudan
            "administrative area 2",  # Juba
            "birth date",  # 2000-06-24
            "estimated birth date",  # TRUE
            "gender",  # FEMALE,
            "marital status",  # MARRIED
            "disability",  # TRUE
            "observed disability",
            "communication disability",
            "hearing disability",  # LOT_DIFFICULTY
            "remembering disability",
            "physical disability",
            "seeing disability",
            "self-care disability",
            "pregnant",  # TRUE
            "relationship to hoh",  # WIFE
            "role",  # PRIMARY
            "work status",  # NOT_PROVIDED
            "sanction list possible match",  # FALSE
            "dedupe in batch status",  # UNIQUE_IN_BATCH
            "dedupe in Pop. status",  # DUPLICATE
            "dedupe in Pop.duplicates",
            "dedupe in Pop. possible duplicates",
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            "household id",
            "country of origin",  # South Sudan
            "administrative area 2",  # Juba
            "household size",  # 4
            "latitude",  # 54,367759
            "longitude",  # 60,964675
            "residence status",  # HOST
            "returnee",  # FALSE
            "status",  # ACTIVE
            "village",  # Mendika
            "females 0-5",  # 0
            "females 0-5 w/ disability",  # 0
            "females 6-11",  # 1
            "females 6-11 w/ disability",  # 1
            "females 12-17",  # 1
            "females 12-17 w/ disability",  # 0
            "females 18-59",  # 1
            "females 18-59 w/ disability",  # 0
            "females 60+",  # 0
            "females 60+ w/ disability",  # 0
            "pregnant females",  # 0
            "males 0-5",  # 0
            "males 0-5 w/ disability",  # 0
            "males 6-11",  # 1
            "males 6-11 w/ disability",  # 1
            "males 12-17",  # 1
            "males 12-17 w/ disability",  # 0
            "males 18-59",  # 1
            "males 18-59 w/ disability",  # 0
            "males 60+",  # 0
            "males 60+ w/ disability",  # 0
            "first registration date",  # 2020-08-25
            "last registration date",  # 2020-08-25
            "organization name enumerator",
        ),
        Report.CASH_PLAN_VERIFICATION: (
            "cash plan ID",  # ANT-21-CSH-00001
            "id",
            "programme",  # Winterization 2020
            "activation date",
            "status",
            "verification method",
            "completion date",
            "sample size",  # 500
            "responded",  # 340
            "received",  # 320
            "received with issues",  # 12
            "not received",  # 8
            "sampling",  # FULL_LIST or RANDOM
            "gender filter",  # FEMALE
            "excluded admin areas",  # Juba
            "age filter",  # {'max': 100, 'min': 0}
        ),
        Report.PAYMENTS: (
            "cash plan ID",  # ANT-21-CSH-00001
            "payment record ID",  # ANT-21-CSH-00001-0000002
            "status",  # Transaction successful
            "currency",
            "delivered quantity",  # 999,00
            "delivery date",  # 2020-11-02 07:50:18+00
            "delivery type",  # deposit to card
            "distribution modality",  # 10K AFN per hh
            "entitlement quantity",  # 1000,00
            "TP ID",
            "TP name",
            "cash or voucher",  # if voucher or e-voucher -> voucher, else -> cash
        ),
        Report.PAYMENT_VERIFICATION: (
            "payment record ID",  # ANT-21-CSH-00001-0000002
            "cash plan ID",  # ANT-21-CSH-00001
            "cash plan verification id",
            "completion date",
            "received amount",  # 30,00
            "status",  # RECEIVED_WITH_ISSUES
            "status date",
        ),
        Report.CASH_PLAN: (
            "program_name",
            "assistance_measurement",
            "assistance_through",
            "business_area_id",
            "ca_hash_id",
            "delivery_type",
            "dispersion_date",
            "down_payment",
            "end_date",
            "funds_commitment",
            "name",
            "program_id",
            "start_date",
            "status",
            "status_date",
            "total_delivered_quantity",
            "total_entitled_quantity",
            "total_entitled_quantity_revised",
            "total_persons_covered",
            "total_persons_covered_revised",
            "total_undelivered_quantity",
            "validation_alerts_count",
            "verification_status",
            "vision_id",  # 54
        ),
        Report.PROGRAM: (
            "business_area_id",
            "administrative_areas_of_implementation",  # Test
            "budget",  # 10000.00
            "cash_plus",  # False
            "description",  # Description goes here
            "end_date",  # 2020-11-17
            "frequency_of_payments",  # REGULAR
            "id",  # e46064c4-d5e2-4990-bb9b-f5cc2dde96f9
            "name",  # Programme 13/10/2020 04:43:28
            "population_goal",  # 50
            "scope",  # UNICEF
            "sector",  # EDUCATION
            "start_date",  # 2020-10-13
            "status",  # ACTIVE
            "total_number_of_households",  # Payment records with delivered amount  > 0 to distinct households
        ),
        Report.INDIVIDUALS_AND_PAYMENT: (
            "admin_area_id",
            "business_area_id",
            "program_name",
            "household_unicef_id",  # HH-20-0000.0368
            "household_country_origin",  # TM
            "birth_date",  # 2000-06-24
            "comms_disability",
            "deduplication_batch_results",  # {"duplicates": [], "possible_duplicates": []}
            "deduplication_golden_record_results",  # {"duplicates": [], "possible_duplicates": []}
            "deduplication_golden_record_status",  # UNIQUE
            "disability",
            "estimated_birth_date",  # False
            "hearing_disability",
            "marital_status",
            "memory_disability",
            "observed_disability",  # NONE
            "physical_disability",
            "pregnant",  # False
            "relationship",  # NON_BENEFICIARY
            "sanction_list_possible_match",  # False
            "seeing_disability",
            "selfcare_disability",
            "sex",  # FEMALE
            "work_status",  # NOT_PROVIDED
            "role",  # PRIMARY
            "currency",
            "delivered_quantity",  # Sum
            "delivered_quantity_usd",
        ),
    }
    OPTIONAL_HEADERS = {Report.HOUSEHOLD_DEMOGRAPHICS: "programme enrolled"}
    ROW_CONTENT_METHODS = {
        Report.INDIVIDUALS: (
            GenerateReportContentHelpers._get_individuals,
            GenerateReportContentHelpers._format_individual_row,
        ),
        Report.HOUSEHOLD_DEMOGRAPHICS: (
            GenerateReportContentHelpers._get_households,
            GenerateReportContentHelpers._format_household_row,
        ),
        Report.CASH_PLAN_VERIFICATION: (
            GenerateReportContentHelpers._get_cash_plan_verifications,
            GenerateReportContentHelpers._format_cash_plan_verification_row,
        ),
        Report.PAYMENTS: (GenerateReportContentHelpers._get_payments, GenerateReportContentHelpers._format_payment_row),
        Report.PAYMENT_VERIFICATION: (
            GenerateReportContentHelpers._get_payment_verifications,
            GenerateReportContentHelpers._format_payment_verification_row,
        ),
        Report.CASH_PLAN: (
            GenerateReportContentHelpers._get_cash_plans,
            GenerateReportContentHelpers._format_cash_plan_row,
        ),
        Report.PROGRAM: (GenerateReportContentHelpers._get_programs, GenerateReportContentHelpers._format_program_row),
        Report.INDIVIDUALS_AND_PAYMENT: (
            GenerateReportContentHelpers._get_payments_for_individuals,
            GenerateReportContentHelpers._format_payments_for_individuals_row,
        ),
    }
    FILTERS_SHEET = "Filters"
    MAX_COL_WIDTH = 50

    def __init__(self, report: Report):
        self.report = report
        self.report_type = report.report_type
        self.business_area = report.business_area

    def _create_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws_report = wb.active
        ws_report.title = f"{self._report_type_to_str()} Report"
        self.wb = wb
        self.ws_report = ws_report
        self.ws_filters = wb.create_sheet(GenerateReportService.FILTERS_SHEET)
        return wb

    def _add_filters_info(self):
        filter_rows = [
            ("Report type", str(self._report_type_to_str())),
            ("Business area", self.business_area.name),
            ("From date", str(self.report.date_from)),
            ("To date", str(self.report.date_to)),
        ]

        if self.report.admin_area.all().exists():
            filter_rows.append(
                (
                    "Administrative area 2",
                    GenerateReportContentHelpers._to_values_list(self.report.admin_area.all(), "title"),
                )
            )
        if self.report.program:
            filter_rows.append(("Program", self.report.program.name))

        for filter_row in filter_rows:
            self.ws_filters.append(filter_row)

    def _add_headers(self):
        headers_row = GenerateReportService.HEADERS[self.report_type]
        self.ws_report.append(headers_row)

    def _add_rows(self) -> int:
        get_row_methods = GenerateReportService.ROW_CONTENT_METHODS[self.report_type]
        all_instances = get_row_methods[0](self.report)
        number_of_columns_based_on_set_headers = len(GenerateReportService.HEADERS[self.report_type])
        col_instances_len = 0
        for instance in all_instances:
            row = get_row_methods[1](instance)
            str_row = self._stringify_all_values(row)
            if len(str_row) > col_instances_len:
                col_instances_len = len(str_row)
            self.ws_report.append(str_row)
        if col_instances_len > number_of_columns_based_on_set_headers:
            # to cover bases when we create extra columns for reverse foreign key instances and we don't know in advance how many columns there will be
            self._add_missing_headers(
                self.ws_report,
                number_of_columns_based_on_set_headers + 1,
                col_instances_len,
                GenerateReportService.OPTIONAL_HEADERS.get(self.report_type, ""),
            )
        return col_instances_len

    def generate_workbook(self) -> openpyxl.Workbook:
        self._create_workbook()
        self._add_filters_info()
        self._add_headers()
        number_of_columns = self._add_rows()
        self._adjust_column_width_from_col(self.ws_filters, 1, 2, 1)
        self._adjust_column_width_from_col(self.ws_report, 1, number_of_columns, 0)
        return self.wb

    def generate_report(self):
        try:
            self.generate_workbook()
            with NamedTemporaryFile() as tmp:
                self.wb.save(tmp.name)
                tmp.seek(0)
                self.report.file.save(
                    f"Report:_{self._report_type_to_str()}_{str(self.report.created_at)}.xlsx", File(tmp), save=False
                )
                self.report.status = Report.COMPLETED
        except Exception as e:
            print("ERROR", e)
            self.report.status = Report.FAILED
        self.report.save()

        if self.report.file:
            self._send_email()

    def _send_email(self):
        # TODO update context when email content is known
        text_body = render_to_string("report.txt", {})
        html_body = render_to_string("report.html", {})
        msg = EmailMultiAlternatives(
            subject="Your report",
            from_email=settings.EMAIL_HOST_USER,
            to=[self.report.created_by.email],
            # cc=[settings.SANCTION_LIST_CC_MAIL],
            body=text_body,
        )
        msg.attach(
            self.report.file.name,
            self.report.file.read(),
            "application/vnd.ms-excel",
        )
        msg.attach_alternative(html_body, "text/html")
        msg.send()

    def _add_missing_headers(self, ws, column_to_start, column_to_finish, label):
        for x in range(column_to_start, column_to_finish + 1):
            col_letter = get_column_letter(x)
            ws[f"{col_letter}1"] = label

    def _adjust_column_width_from_col(self, ws, min_col, max_col, min_row):

        column_widths = []

        for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):

            for cell in col:
                value = cell.value
                if value is not None:

                    if isinstance(value, str) is False:
                        value = str(value)

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

        for i, width in enumerate(column_widths):
            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            value = GenerateReportService.MAX_COL_WIDTH if value > GenerateReportService.MAX_COL_WIDTH else value
            ws.column_dimensions[col_name].width = value

    def _report_type_to_str(self) -> str:
        return [name for value, name in Report.REPORT_TYPES if value == self.report_type][0]

    def _stringify_all_values(self, row: tuple) -> tuple:
        str_row = []
        for value in row:
            str_row.append(str(value if value is not None else ""))
        return tuple(str_row)
